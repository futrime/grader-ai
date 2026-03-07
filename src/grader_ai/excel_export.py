"""Export grading results to Excel (.xls or .xlsx) files."""

import logging
from pathlib import Path

from grader_ai.core import Report

logger = logging.getLogger(__name__)

# Column header names (Chinese) - matches Assignment 0_学生名单列表.xls format
ID_HEADERS = ("学号", "学生编号", "student_id", "ID")
TOTAL_HEADERS = ("成绩（录入项）", "成绩（录入区）", "总分", "total", "Total")
COMMENT_HEADERS = ("评语（录入项）", "评语（录入区）", "评语", "反馈", "comments", "feedback", "Comment")


def _extract_student_id(submission_file: str) -> str:
    """Extract student ID from submission filename (e.g., 2024010684_程思元_5819.zip)."""
    stem = Path(submission_file).stem
    parts = stem.split("_")
    return parts[0] if parts else stem


def _find_column_index(header_row: list, candidates: tuple[str, ...]) -> int | None:
    """Find column index: exact match first, then contains (for 成绩（录入区）etc)."""
    for i, cell in enumerate(header_row):
        val = str(cell).strip() if cell is not None else ""
        for cand in candidates:
            if val == cand or cand in val:
                return i
    return None


def _update_excel_xlsx(
    excel_path: Path,
    reports: list[Report],
) -> None:
    """Update .xlsx file using openpyxl."""
    import openpyxl

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    header_row = [cell.value for cell in ws[1]]
    id_col = _find_column_index(header_row, ID_HEADERS)
    if id_col is None:
        raise ValueError(
            "Excel must have a column header for student ID (学号, 学生编号, etc.)"
        )
    total_col = _find_column_index(header_row, TOTAL_HEADERS)
    comment_col = _find_column_index(header_row, COMMENT_HEADERS)

    reports_by_id = {_extract_student_id(r.submission_file): r for r in reports}

    for row_idx in range(2, ws.max_row + 1):
        row_id = ws.cell(row=row_idx, column=id_col + 1).value
        if row_id is None:
            continue
        sid = str(row_id).strip()
        report = reports_by_id.get(sid)
        if report is None:
            continue

        if report.error:
            total_score = 0
            feedback_text = f"批阅失败: {report.error}"
        else:
            total_score = sum(gr.score for gr in report.grade_results)
            deducted = [
                f"Q{i+1} -{gr.credits - gr.score}"
                for i, gr in enumerate(report.grade_results)
                if gr.score < gr.credits
            ]
            feedback_text = "; ".join(deducted) if deducted else ""

        if total_col is not None:
            ws.cell(row=row_idx, column=total_col + 1, value=total_score)
        if comment_col is not None:
            ws.cell(row=row_idx, column=comment_col + 1, value=feedback_text)

    wb.save(excel_path)
    logger.info("Wrote grades to %s", excel_path)


def _update_excel_xls(
    excel_path: Path,
    reports: list[Report],
) -> None:
    """Update .xls file using xlrd (read) and xlwt (write)."""
    import xlrd
    import xlwt

    rb = xlrd.open_workbook(str(excel_path), formatting_info=False)
    rs = rb.sheet_by_index(0)

    header_row = [rs.cell_value(0, c) for c in range(rs.ncols)]
    id_col = _find_column_index(header_row, ID_HEADERS)
    if id_col is None:
        raise ValueError(
            "Excel must have a column header for student ID (学号, 学生编号, etc.)"
        )
    total_col = _find_column_index(header_row, TOTAL_HEADERS)
    comment_col = _find_column_index(header_row, COMMENT_HEADERS)

    reports_by_id = {_extract_student_id(r.submission_file): r for r in reports}

    wb = xlwt.Workbook()
    ws = wb.add_sheet(rs.name, cell_overwrite_ok=True)

    for c in range(rs.ncols):
        ws.write(0, c, rs.cell_value(0, c))

    for r in range(1, rs.nrows):
        for c in range(rs.ncols):
            val = rs.cell_value(r, c)
            if isinstance(val, float) and val == int(val):
                val = int(val)
            ws.write(r, c, val)

        if id_col is not None:
            row_id = rs.cell_value(r, id_col)
            sid = str(row_id).strip()
            report = reports_by_id.get(sid)
            if report is not None:
                if report.error:
                    total_score = 0
                    feedback_text = f"批阅失败: {report.error}"
                else:
                    total_score = sum(gr.score for gr in report.grade_results)
                    deducted = [
                        f"Q{i+1} -{gr.credits - gr.score}"
                        for i, gr in enumerate(report.grade_results)
                        if gr.score < gr.credits
                    ]
                    feedback_text = "; ".join(deducted) if deducted else ""

                if total_col is not None:
                    ws.write(r, total_col, total_score)
                if comment_col is not None:
                    ws.write(r, comment_col, feedback_text)

    wb.save(str(excel_path))
    logger.info("Wrote grades to %s", excel_path)


def export_to_excel(excel_path: Path, reports: list[Report]) -> None:
    """Read Excel file, update rows with grading results, overwrite file."""
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")
    suffix = excel_path.suffix.lower()
    if suffix in (".xlsx",):
        _update_excel_xlsx(excel_path, reports)
    elif suffix in (".xls",):
        _update_excel_xls(excel_path, reports)
    else:
        raise ValueError(
            f"Unsupported Excel format: {suffix}. Use .xls or .xlsx"
        )
